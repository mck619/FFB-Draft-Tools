
from bs4 import BeautifulSoup
import requests
import string
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter

def get_fp_table(week, position):
    '''
    :param week: int value of current nfl week
    :param position: qb, rb, wr or te
    :return: list of projections for that week and position from fantasypros.com
    '''

    url = 'http://www.fantasypros.com/nfl/projections/'+position+'.php?week='+str(week)

    r = requests.get(url)
    data = r.text
    soup = BeautifulSoup(data, "html.parser")
    tb = soup.find('table', attrs={'id':'data'})
    col_titles = col_list(position)
    players = []

    trs = tb.find_all(class_ = 'mpb-available')

    for row in trs:
        dps =  row.find_all('td')
        i = 0
        row_dict = {}
        for data_point in dps:
            try:
                row_dict[col_titles[i]] = float(data_point.text)
            except:
                row_dict[col_titles[i]] = str(data_point.find('a').text)

            i += 1
        players.append(row_dict)

    return {'site': 'FantasyPros','position': position, 'week': week, 'projections': players}


def col_list(position):
    if position == 'qb':
        col_titles = ['Name', 'pass_att', 'pass_cmp', 'pass_yds', 'pass_tds',
                      'pass_int', 'rush_att', 'rush_yds', 'rush_tds', 'fl', 'fpts']
    elif position == 'wr' or 'rb':
        col_titles = ['Name', 'rush_att', 'rush_yds', 'rush_tds', 'rec', 'rec_yds', 'rec_tds', 'fl', 'fpts']
    else:
        col_titles = ['Name', 'rec', 'rec_yds', 'rec_tds', 'fl', 'fpts']
    return col_titles

def write_ws(wb, table):
    ws_cols = list(string.ascii_uppercase)

    row = 1
    col = 0
    ws = wb.create_sheet()
    ws.title = table['site']+'_week_'+ str(table['week'])+"_"+table['position']
    col_titles = col_list(table['position'])

    for title in col_titles:
        ws[ws_cols[col] + str(row)] = title
        col+=1
    col = 0
    row +=1

    for proj in table['projections']:
        for col_t in col_titles:
            ws[ws_cols[col] + str(row)] = proj[col_t]
            col += 1
        row += 1
        col = 0



table = get_fp_table(1,'rb')
wb = Workbook()
write_ws(wb,table)
wb.save('test.xlsx')










